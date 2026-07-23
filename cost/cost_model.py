#!/usr/bin/env python3
"""
Modelo de costo — TECHO de máxima complejidad del framework SDD de Valtx.

Objetivo: dimensionar el gasto MENSUAL a distintas escalas (features/mes) para la
conversación de recursos con la GG. Es un TECHO deliberado (pipeline completo, gates
todos activos, agentes en worktrees, modelo Opus en la fase dominante). Desde este
techo se refina hacia abajo.

Todos los supuestos están arriba y son editables. Precios en USD, aprox. — VERIFICAR
tarifas vigentes antes de presentar. Ejecuta:  python cost/cost_model.py
Escribe además un Excel con el desglose.
"""
import os, math

# ─────────────────────────────────────────────────────────────────────────────
# 1) PRECIOS LLM (USD por 1,000,000 tokens)  [in, out]   — ajustar a tarifa vigente
# ─────────────────────────────────────────────────────────────────────────────
PRICES = {
    "haiku":  (0.80, 4.00),
    "sonnet": (3.00, 15.00),
    "opus":   (15.00, 75.00),
    "codex":  (3.00, 15.00),   # asimilado a tier medio (GPT-class)
}
CACHED_INPUT_FACTOR = 0.10     # lectura desde prompt-cache ≈ 10% del precio de input

# ─────────────────────────────────────────────────────────────────────────────
# 2) CONSUMO POR FEATURE — pipeline COMPLETO (techo). Tokens por etapa.
#    iter = veces que se repite (incluye regenerar-con-referencias ~1.4).
#    cache = fracción del input servida desde caché (contexto estable reutilizado).
# ─────────────────────────────────────────────────────────────────────────────
STAGES = [
    # etapa,                 modelo,    in,      out,    iter, cache
    ("Refinamiento prompt",  "haiku",   5_000,   2_000,  1.0,  0.0),
    ("Spec + EARS",          "sonnet",  15_000,  8_000,  1.4,  0.3),
    ("Clarify",              "sonnet",  10_000,  4_000,  1.2,  0.5),
    ("Plan / Tech Spec",     "sonnet",  20_000,  10_000, 1.4,  0.4),
    ("Tasks",                "sonnet",  12_000,  6_000,  1.2,  0.5),
    ("Gherkin scenarios",    "sonnet",  10_000,  5_000,  1.2,  0.4),
    # Fase DOMINANTE: desarrollo agéntico. 8 task-issues, loop con retrieval+gen.
    ("Code dev (8 tasks·Opus)", "opus", 480_000, 160_000, 1.4, 0.7),
    ("Gherkin step defs / tests", "codex", 40_000, 20_000, 1.3, 0.5),
    ("Review / verify",      "sonnet",  25_000,  8_000,  1.4,  0.6),
]

# ─────────────────────────────────────────────────────────────────────────────
# 3) PLATAFORMAS (fijo por asiento/mes) y throughput
# ─────────────────────────────────────────────────────────────────────────────
SEAT_COST = {           # USD por usuario/mes (planes de gama alta = techo)
    "github_enterprise": 21.0,
    "linear_business":   14.0,
    "notion_business":   20.0,   # incluye margen para add-ons
}
SEAT_TOTAL = sum(SEAT_COST.values())
FEATURES_PER_SEAT = 12          # features/mes que sostiene 1 asiento (dev agéntico)
MIN_SEATS = 3                    # piso operativo (dev + PMO + legal/gobernanza)

# ─────────────────────────────────────────────────────────────────────────────
# 4) INFRAESTRUCTURA (variable por feature + fijo)
# ─────────────────────────────────────────────────────────────────────────────
CI_MIN_PER_FEATURE = 300        # min de runners (agentes) por feature
CI_COST_PER_MIN     = 0.008     # GitHub Actions linux (hosted)
VERTEX_PER_FEATURE  = 1.20      # búsquedas + indexado KB por feature
PREVIEW_PER_FEATURE = 0.80      # entorno efímero por feature
INFRA_FIXED_MONTH   = 400.0     # KB/Vertex base, storage, observabilidad, secret mgr

SCALES = [20, 50, 150, 500, 1500]   # features/mes


def eff_input_price(model, cache):
    pin = PRICES[model][0]
    return pin * (1 - cache * (1 - CACHED_INPUT_FACTOR))


def stage_cost(model, tin, tout, it, cache):
    c_in = tin / 1e6 * eff_input_price(model, cache)
    c_out = tout / 1e6 * PRICES[model][1]
    return (c_in + c_out) * it


def per_feature_breakdown(stages=STAGES):
    rows, total = [], 0.0
    for name, model, tin, tout, it, cache in stages:
        c = stage_cost(model, tin, tout, it, cache)
        rows.append((name, model, int(tin*it), int(tout*it), round(c, 3)))
        total += c
    return rows, total


def optimized_stages():
    """Piso: fase dominante a Sonnet, más caché, menos iteraciones. Mismo pipeline."""
    opt = []
    for name, model, tin, tout, it, cache in STAGES:
        if name.startswith("Code dev"):
            opt.append((name.replace("Opus", "Sonnet"), "sonnet", tin, tout, 1.2, 0.85))
        else:
            opt.append((name, model, tin, tout, max(1.0, it-0.1), min(0.9, cache+0.15)))
    return opt


def monthly(features, per_feat_llm):
    seats = max(MIN_SEATS, math.ceil(features / FEATURES_PER_SEAT))
    llm = per_feat_llm * features
    platforms = seats * SEAT_TOTAL
    infra = (CI_MIN_PER_FEATURE * CI_COST_PER_MIN + VERTEX_PER_FEATURE
             + PREVIEW_PER_FEATURE) * features + INFRA_FIXED_MONTH
    return dict(features=features, seats=seats, llm=llm, platforms=platforms,
                infra=infra, total=llm + platforms + infra)


def fmt(x):
    return f"${x:,.0f}"


def main():
    rows, per_feat = per_feature_breakdown()
    print("\n" + "═"*78)
    print(" TECHO DE COSTO · desglose POR FEATURE (pipeline completo, máx. complejidad)")
    print("═"*78)
    print(f" {'Etapa':28} {'Modelo':7} {'tok_in':>9} {'tok_out':>9} {'USD':>8}")
    for name, model, tin, tout, c in rows:
        print(f" {name:28} {model:7} {tin:>9,} {tout:>9,} {c:>8.2f}")
    print("─"*78)
    print(f" {'COSTO LLM POR FEATURE':28} {'':7} {'':>9} {'':>9} {per_feat:>8.2f}")

    print("\n" + "═"*78)
    print(" TECHO MENSUAL POR ESCALA (USD/mes)")
    print("═"*78)
    print(f" {'features/mes':>12} {'seats':>6} {'LLM':>12} {'plataformas':>12} {'infra':>10} {'TOTAL':>12} {'$/feat':>9}")
    scale_rows = []
    for f in SCALES:
        m = monthly(f, per_feat)
        perf = m["total"]/f
        scale_rows.append(m)
        print(f" {m['features']:>12,} {m['seats']:>6} {fmt(m['llm']):>12} "
              f"{fmt(m['platforms']):>12} {fmt(m['infra']):>10} {fmt(m['total']):>12} {perf:>8.0f}")
    # Escenario optimizado (piso) para contraste
    _, per_feat_opt = per_feature_breakdown(optimized_stages())
    print("═"*78)
    print(f" ESCENARIO OPTIMIZADO (piso): LLM/feature ${per_feat_opt:.2f}  "
          f"(vs techo ${per_feat:.2f} = -{100-100*per_feat_opt/per_feat:.0f}%)")
    print(f"   Mismo pipeline y gates; solo fase dominante a Sonnet + más caché.")
    print(f" {'features/mes':>12} {'TECHO/mes':>12} {'PISO/mes':>12}")
    for f in SCALES:
        mt = monthly(f, per_feat); mo = monthly(f, per_feat_opt)
        print(f" {f:>12,} {fmt(mt['total']):>12} {fmt(mo['total']):>12}")
    print("═"*78)
    print(" El TECHO es para pedir presupuesto; el PISO es a dónde optimizas operando.\n")

    write_excel(rows, per_feat, scale_rows, per_feat_opt)


def write_excel(rows, per_feat, scale_rows, per_feat_opt=None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("(openpyxl no disponible; se omite Excel)"); return
    out = os.path.join(os.path.dirname(__file__), "..", "..", "Valtx_SDD_Modelo_Costo.xlsx")
    wb = Workbook()
    H = Font(bold=True, color="FFFFFF"); HF = PatternFill("solid", fgColor="00235F")
    T = Font(bold=True, color="00235F")

    ws = wb.active; ws.title = "Por_Feature"
    ws.append(["Etapa", "Modelo", "tokens_in", "tokens_out", "USD/feature"])
    for c in ws[1]: c.font = H; c.fill = HF
    for r in rows: ws.append(list(r))
    ws.append(["COSTO LLM POR FEATURE", "", "", "", round(per_feat, 2)])
    ws[ws.max_row][0].font = T; ws[ws.max_row][4].font = T
    for col, w in zip("ABCDE", [26, 9, 12, 12, 13]): ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Techo_Mensual")
    ws2.append(["features/mes", "seats", "LLM", "Plataformas", "Infra", "TOTAL/mes", "USD/feature", "TOTAL/año"])
    for c in ws2[1]: c.font = H; c.fill = HF
    for m in scale_rows:
        ws2.append([m["features"], m["seats"], round(m["llm"]), round(m["platforms"]),
                    round(m["infra"]), round(m["total"]), round(m["total"]/m["features"], 1),
                    round(m["total"]*12)])
    for col, w in zip("ABCDEFGH", [13, 7, 12, 13, 10, 12, 12, 12]): ws2.column_dimensions[col].width = w
    for row in ws2.iter_rows(min_row=2, min_col=3, max_col=6):
        for c in row: c.number_format = '#,##0'

    ws3 = wb.create_sheet("Supuestos")
    assump = [
        ["SUPUESTOS (editar y recalcular)", ""],
        ["Precio Opus in/out (USD/1M)", f"{PRICES['opus'][0]} / {PRICES['opus'][1]}"],
        ["Precio Sonnet in/out (USD/1M)", f"{PRICES['sonnet'][0]} / {PRICES['sonnet'][1]}"],
        ["Precio Haiku in/out (USD/1M)", f"{PRICES['haiku'][0]} / {PRICES['haiku'][1]}"],
        ["Factor lectura caché", CACHED_INPUT_FACTOR],
        ["Costo por asiento/mes (3 plataformas)", SEAT_TOTAL],
        ["Features por asiento/mes", FEATURES_PER_SEAT],
        ["Asientos mínimos", MIN_SEATS],
        ["CI min/feature x USD/min", f"{CI_MIN_PER_FEATURE} x {CI_COST_PER_MIN}"],
        ["Vertex + preview por feature (USD)", VERTEX_PER_FEATURE + PREVIEW_PER_FEATURE],
        ["Infra fija/mes (USD)", INFRA_FIXED_MONTH],
        ["Tasks por feature (fase dominante)", 8],
        ["", ""],
        ["NOTA", "Techo de máx. complejidad. Precios aprox.; verificar tarifas vigentes."],
    ]
    for r in assump: ws3.append(r)
    ws3["A1"].font = T
    ws3.column_dimensions["A"].width = 40; ws3.column_dimensions["B"].width = 50

    if per_feat_opt is not None:
        ws4 = wb.create_sheet("Techo_vs_Piso")
        ws4.append(["features/mes", "TECHO/mes", "PISO/mes", "Ahorro", "TECHO/año", "PISO/año"])
        for c in ws4[1]: c.font = H; c.fill = HF
        for f in SCALES:
            mt = monthly(f, per_feat); mo = monthly(f, per_feat_opt)
            ws4.append([f, round(mt["total"]), round(mo["total"]),
                        round(mt["total"]-mo["total"]), round(mt["total"]*12), round(mo["total"]*12)])
        for col, w in zip("ABCDEF", [13, 12, 12, 12, 12, 12]): ws4.column_dimensions[col].width = w
        for row in ws4.iter_rows(min_row=2):
            for c in row: c.number_format = '#,##0'

    wb.save(out)
    print(f" Excel escrito: {os.path.abspath(out)}")


if __name__ == "__main__":
    main()
